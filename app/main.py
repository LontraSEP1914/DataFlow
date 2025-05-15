# ... (importações e classes LogLevel, ConsolidationWorker como antes) ...
import sys
import os
import glob
import time 
import polars as pl
import openpyxl 
import xlrd
import json

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLineEdit, QLabel, QListWidget, QListWidgetItem,
    QComboBox, QProgressBar, QTextEdit, QFileDialog, QTabWidget, 
    QTableView, QDialogButtonBox, QTableWidget, QDialog, QTableWidgetItem,
    QCheckBox, QHeaderView
)
from PySide6.QtCore import Qt, QThread, Signal , QAbstractTableModel
from enum import Enum

# Definir os tipos de dados que o usuário pode escolher
DATA_TYPES_OPTIONS = ["Automático/String", "Inteiro", "Decimal (Float)", "Data", "Booleano"]
# Mapeamento para tipos Polars (pode ser um dict global ou dentro do worker)
TYPE_STRING_TO_POLARS = {
    "Automático/String": pl.String,
    "Inteiro": pl.Int64,
    "Decimal (Float)": pl.Float64,
    "Data": pl.Date, # ou pl.Datetime se precisar de hora
    "Booleano": pl.Boolean
}

CONFIG_FILE_NAME = "config_consolidador.json" # Nome do arquivo de configuração

# ... (LogLevel e ConsolidationWorker aqui) ...
class LogLevel(Enum):
    INFO = "[INFO]"
    WARNING = "[AVISO]"
    ERROR = "[ERRO]"
    SUCCESS = "[SUCESSO]"

class ConsolidationWorker(QThread):
    progress_updated = Signal(int) 
    log_message = Signal(str, LogLevel) 
    finished = Signal(bool, str) 

    def __init__(self, files_to_process, output_path, output_format, header_mapping):
        super().__init__()
        self.files_to_process = files_to_process 
        self.output_path = output_path
        self.output_format = output_format
        self.header_mapping = header_mapping
        self.is_running = True

    def run(self):
        try:
            self.log_message.emit("Iniciando processo de consolidação...", LogLevel.INFO)
            all_dataframes_processed = [] 
            
            total_items = 0
            for _, sheets_to_process_for_file in self.files_to_process:
                if sheets_to_process_for_file is None: total_items += 1
                else: total_items += len(sheets_to_process_for_file)
            if total_items == 0:
                self.log_message.emit("Nenhum item válido para processar.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum item para processar.")
                return
            processed_items = 0

            for file_path, selected_sheets in self.files_to_process:
                if not self.is_running: break 
                file_name = os.path.basename(file_path)
                sheets_to_iterate = selected_sheets if selected_sheets is not None else [None]

                for sheet_name in sheets_to_iterate:
                    if not self.is_running: break 
                    current_item_description = f"'{file_name}'" + (f" - Aba: '{sheet_name}'" if sheet_name else "")
                    
                    try:
                        df_original = None
                        if sheet_name is None: 
                            df_original = pl.read_csv(source=file_path, infer_schema_length=1000, try_parse_dates=True)
                        else: 
                            df_original = pl.read_excel(source=file_path, sheet_name=sheet_name)

                        if df_original is None or df_original.height == 0:
                            self.log_message.emit(f"Dados vazios ou erro ao ler {current_item_description}. Pulando.", LogLevel.WARNING)
                            processed_items += 1 
                            continue

                        # --- 1. Aplicar Mapeamento de Nomes e Filtro de Colunas ---
                        df_intermediate = df_original 
                        if self.header_mapping:
                            selected_columns_expressions = []
                            rename_map_for_select = {} # Para renomear DURANTE o select
                            
                            # Primeiro, determinar quais colunas originais manter e seus nomes finais
                            # Estrutura auxiliar: {original_col: final_col_name_if_kept}
                            cols_to_keep_and_their_final_names = {}

                            for original_col_name_from_file in df_original.columns:
                                mapping_info = self.header_mapping.get(original_col_name_from_file)
                                if mapping_info: # Coluna do arquivo está no mapeamento
                                    if mapping_info.get("include", False):
                                        final_name = mapping_info.get("final_name", original_col_name_from_file)
                                        cols_to_keep_and_their_final_names[original_col_name_from_file] = final_name
                                else: # Coluna do arquivo NÃO está no mapeamento -> incluir com nome original
                                    cols_to_keep_and_their_final_names[original_col_name_from_file] = original_col_name_from_file
                            
                            if not cols_to_keep_and_their_final_names:
                                self.log_message.emit(f"Nenhuma coluna selecionada/mapeada para {current_item_description}. Pulando.", LogLevel.WARNING)
                                processed_items += 1; continue

                            # Construir expressões de select com alias para renomeação
                            select_expressions = [
                                pl.col(orig_name).alias(final_name) 
                                for orig_name, final_name in cols_to_keep_and_their_final_names.items()
                                if orig_name in df_original.columns # Garantir que a coluna original existe no df
                            ]
                            
                            if not select_expressions: # Se após filtro, nenhuma coluna do df atual sobrou
                                self.log_message.emit(f"Nenhuma coluna do arquivo {current_item_description} corresponde ao mapeamento. Pulando.", LogLevel.WARNING)
                                processed_items +=1; continue

                            df_intermediate = df_original.select(select_expressions)
                        
                        if df_intermediate.width == 0:
                            self.log_message.emit(f"Nenhuma coluna restante em {current_item_description} após mapeamento de nomes. Pulando.", LogLevel.WARNING)
                            processed_items += 1; continue
                        
                        # --- 2. Aplicar Tipagem Especificada pelo Usuário ---
                        df_typed = df_intermediate
                        if self.header_mapping:
                            # Precisamos iterar sobre as colunas FINAIS do df_intermediate
                            # e encontrar a regra de tipagem correspondente no header_mapping
                            # (que é chaveado pelo nome ORIGINAL).
                            
                            # Criar um mapa reverso: {final_name: original_name_info_from_mapping}
                            # ou mais direto: {final_name: type_str_from_mapping}
                            final_name_to_type_str = {}
                            for original_h, map_details in self.header_mapping.items():
                                if map_details.get("include"):
                                    final_name_to_type_str[map_details.get("final_name", original_h)] = map_details.get("type_str")

                            casting_expressions = []
                            for final_col_name in df_typed.columns: # Iterar sobre colunas já mapeadas/renomeadas
                                type_str = final_name_to_type_str.get(final_col_name)
                                
                                if type_str and type_str != DATA_TYPES_OPTIONS[0]: # Se não for "Automático/String"
                                    polars_type = TYPE_STRING_TO_POLARS.get(type_str)
                                    if polars_type:
                                        self.log_message.emit(f"Convertendo coluna '{final_col_name}' para {type_str} em {current_item_description}", LogLevel.INFO)
                                        casting_expressions.append(pl.col(final_col_name).cast(polars_type, strict=False))
                                    else: # Tipo string não encontrado no mapa (não deve acontecer)
                                        casting_expressions.append(pl.col(final_col_name)) # Manter como está
                                else: # "Automático/String" ou tipo não mapeado
                                    casting_expressions.append(pl.col(final_col_name)) # Manter como está
                            
                            if casting_expressions: # Se houver alguma expressão (sempre haverá se df_typed não for vazio)
                                df_typed = df_typed.select(casting_expressions) # .select() é mais seguro que .with_columns() para recriar
                        
                        all_dataframes_processed.append(df_typed)

                    except Exception as e:
                        self.log_message.emit(f"Erro ao processar (ler/mapear/tipar) {current_item_description}: {e}", LogLevel.ERROR)
                    
                    processed_items += 1
                    progress = int((processed_items / total_items) * 100) if total_items > 0 else 0
                    self.progress_updated.emit(progress)

                if not self.is_running: break 

            # ... (Resto do código: checagem de cancelamento, se all_dataframes_processed está vazio, 
            #      harmonização de tipos (que agora opera em df_typed), concatenação final, salvar) ...
            # A lógica de harmonização de tipos que tínhamos (numérico vs string -> string)
            # ainda é útil como uma segunda passagem após a tentativa de cast do usuário,
            # especialmente se strict=False no cast do usuário resultou em muitos nulos e
            # a coluna ainda tem tipos mistos que o concat não lidaria bem.
            # Ou, podemos confiar que o cast com strict=False já deixou os tipos prontos
            # para concatenação (ex: se um Int falha para Date, vira null, e a coluna é Date).
            # Por simplicidade, vamos manter a harmonização por enquanto.

            if not self.is_running:
                 self.log_message.emit("Consolidação cancelada.", LogLevel.WARNING)
                 self.finished.emit(False, "Cancelado"); return

            if not all_dataframes_processed:
                self.log_message.emit("Nenhum dado após processamento.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum dado processado."); return

            # --- Harmonização de Tipos (Pós-Tipagem do Usuário) ---
            self.log_message.emit("Harmonizando tipos (2ª passagem) entre arquivos processados...", LogLevel.INFO)
            # ... (código da harmonização idêntico ao anterior, operando sobre all_dataframes_processed) ...
            # (cole a seção de harmonização aqui)
            column_types_map = {} 
            all_column_names = set()
            for i, df in enumerate(all_dataframes_processed):
                 for col_name, dtype in df.schema.items():
                     all_column_names.add(col_name)
                     if col_name not in column_types_map: column_types_map[col_name] = {}
                     column_types_map[col_name][i] = dtype
            
            harmonized_dataframes_final_pass = [] # Nova lista para a segunda passagem
            for i, df_to_harmonize in enumerate(all_dataframes_processed):
                 df_modified_second_pass = df_to_harmonize 
                 cols_to_cast_to_string_final = []
                 for col_name in df_modified_second_pass.columns:
                     if col_name in column_types_map:
                         types_for_this_col = [dtype_info for df_idx, dtype_info in column_types_map[col_name].items()]
                         is_numeric_type_present = any(t.is_numeric() for t in types_for_this_col)
                         is_string_type_present = any(t == pl.String or t == pl.Utf8 for t in types_for_this_col)
                         # Adicionar checagem para Date/Datetime vs String também, pois podem conflitar
                         is_temporal_type_present = any(t.is_temporal() for t in types_for_this_col)

                         if (is_numeric_type_present and is_string_type_present) or \
                            (is_temporal_type_present and is_string_type_present):
                             cols_to_cast_to_string_final.append(col_name)
                 
                 if cols_to_cast_to_string_final:
                    select_expressions_final = []
                    for col_current_name in df_modified_second_pass.columns:
                        if col_current_name in cols_to_cast_to_string_final:
                            select_expressions_final.append(pl.col(col_current_name).cast(pl.String).alias(col_current_name))
                        else:
                            select_expressions_final.append(pl.col(col_current_name)) 
                    if select_expressions_final: df_modified_second_pass = df_modified_second_pass.select(select_expressions_final)
                 harmonized_dataframes_final_pass.append(df_modified_second_pass)
            
            final_dataframes_to_concat = harmonized_dataframes_final_pass
            # --- Fim Harmonização (2ª passagem) ---


            self.log_message.emit("Concatenando dados processados...", LogLevel.INFO)
            # ... (resto: concatenação e salvamento como antes) ...
            try:
                consolidated_df = pl.concat(final_dataframes_to_concat, how="diagonal") 
            except Exception as e: 
                 self.log_message.emit(f"Erro concatenação final: {e}", LogLevel.ERROR)
                 self.finished.emit(False, f"Erro concatenação: {e}"); return
            
            self.log_message.emit(f"Salvando: {self.output_path}", LogLevel.INFO)
            if self.output_format == "XLSX": consolidated_df.write_excel(self.output_path)
            elif self.output_format == "CSV": consolidated_df.write_csv(self.output_path)

            self.progress_updated.emit(100)
            self.log_message.emit(f"Concluído! Salvo em: {self.output_path}", LogLevel.SUCCESS)
            self.finished.emit(True, f"Salvo em: {self.output_path}")

        except Exception as e:
            self.log_message.emit(f"Erro inesperado consolidação: {e}", LogLevel.ERROR)
            self.finished.emit(False, f"Erro: {e}")

    def stop(self): # stop() permanece o mesmo
        self.is_running = False
        self.log_message.emit("Tentativa de parada da consolidação solicitada...", LogLevel.INFO)

class PolarsTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data if data is not None else pl.DataFrame()

    def rowCount(self, parent=None):
        return self._data.height

    def columnCount(self, parent=None):
        return self._data.width

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.DisplayRole:
            try:
                # Polars pode retornar um único valor ou uma Series.
                # Se for Series (ao pegar uma célula), pegue o primeiro valor.
                value = self._data[index.row(), index.column()]
                if isinstance(value, pl.Series): # Acesso a célula pode retornar Series de 1 elemento
                    return str(value[0]) if value.len() > 0 else ""
                return str(value) # Converte para string para exibição
            except Exception:
                return "" # Em caso de erro ao acessar, retorna string vazia
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal: # Cabeçalhos das colunas
                return str(self._data.columns[section])
            if orientation == Qt.Vertical: # Cabeçalhos das linhas (números)
                return str(section + 1)
        return None

    def load_data(self, new_data: pl.DataFrame):
        self.beginResetModel()
        self._data = new_data if new_data is not None else pl.DataFrame()
        self.endResetModel()

    def clear_data(self):
        self.load_data(pl.DataFrame())

class HeaderMappingDialog(QDialog):
    def __init__(self, unique_headers, parent=None, existing_mapping=None):
        super().__init__(parent)
        self.setWindowTitle("Mapeamento e Tipagem de Cabeçalhos")
        self.setMinimumSize(750, 450) # Aumentar um pouco para a nova coluna

        self.unique_headers = sorted(list(unique_headers))
        self.mapping = existing_mapping if existing_mapping is not None else {}

        layout = QVBoxLayout(self)
        description_label = QLabel(
            "Ajuste os nomes finais, selecione a inclusão e defina o tipo de dados para cada coluna."
        )
        layout.addWidget(description_label)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(4)
        self.table_widget.setHorizontalHeaderLabels(["Cabeçalho Original", "Cabeçalho Final", 'Tipo de Dados', "Incluir?"])
        self.table_widget.setRowCount(len(self.unique_headers))

        for row, header_original in enumerate(self.unique_headers):
            # Cabeçalho Original (não editável)
            item_original = QTableWidgetItem(header_original)
            item_original.setFlags(item_original.flags() & ~Qt.ItemIsEditable) # Remove flag de edição
            self.table_widget.setItem(row, 0, item_original)

            # Cabeçalho Final (QLineEdit)
            final_name_edit = QLineEdit()
            # Pré-preencher com mapeamento existente ou com o original
            final_name = self.mapping.get(header_original, {}).get("final_name", header_original)
            final_name_edit.setText(final_name)
            self.table_widget.setCellWidget(row, 1, final_name_edit)

            # Tipo de Dados (QComboBox) <--- NOVA COLUNA
            type_combo = QComboBox()
            type_combo.addItems(DATA_TYPES_OPTIONS)
            saved_type_str = self.mapping.get(header_original, {}).get("type_str", DATA_TYPES_OPTIONS[0]) # Default "Automático/String"
            if saved_type_str in DATA_TYPES_OPTIONS:
                type_combo.setCurrentText(saved_type_str)
            else: # Se o tipo salvo não for válido, default para o primeiro
                 type_combo.setCurrentIndex(0)
            self.table_widget.setCellWidget(row, 2, type_combo)

            # Incluir? (QCheckBox)
            checkbox_widget = QWidget() # Widget para centralizar o checkbox
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox = QCheckBox()
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(0,0,0,0)
            checkbox_widget.setLayout(checkbox_layout)
            
            # Pré-preencher com mapeamento existente ou True por padrão
            is_included = self.mapping.get(header_original, {}).get("include", True)
            checkbox.setChecked(is_included)
            self.table_widget.setCellWidget(row, 3, checkbox_widget)
            
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents) 
        layout.addWidget(self.table_widget)

        # Botões OK e Cancelar
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_mapping(self):
        """Retorna o mapeamento definido pelo usuário."""
        updated_mapping = {}
        for row in range(self.table_widget.rowCount()):
            original_header_item = self.table_widget.item(row, 0)
            if not original_header_item: continue # Segurança se o item não existir
            original_header = original_header_item.text()

            final_name_widget = self.table_widget.cellWidget(row, 1) # Pega o QLineEdit
            final_name = original_header # Default
            if isinstance(final_name_widget, QLineEdit): # Verifica se realmente é um QLineEdit
                final_name = final_name_widget.text().strip()
            
            type_combo_widget = self.table_widget.cellWidget(row, 2) # <--- LER O COMBOBOX
            selected_type_str = DATA_TYPES_OPTIONS[0] # Default
            if isinstance(type_combo_widget, QComboBox):
                selected_type_str = type_combo_widget.currentText()
            
            checkbox_widget_container = self.table_widget.cellWidget(row, 3) # Pega o QWidget container
            include = True # Default
            if checkbox_widget_container:
                # Encontrar o QCheckBox dentro do layout do QWidget container
                layout = checkbox_widget_container.layout()
                if layout and layout.count() > 0:
                    checkbox = layout.itemAt(0).widget()
                    if isinstance(checkbox, QCheckBox): # Verifica se é um QCheckBox
                        include = checkbox.isChecked()

            # Se o usuário apagar o nome final, usar o original como fallback
            if not final_name: 
                final_name = original_header
                # Opcional: resetar o QLineEdit para o nome original se estiver vazio?
                # if isinstance(final_name_widget, QLineEdit):
                #     final_name_widget.setText(final_name)

            updated_mapping[original_header] = {
                "final_name": final_name,
                "type_str": selected_type_str, # Armazenar a string do tipo
                "include": include
            }
        return updated_mapping

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.header_mapping = {}
        self.setWindowTitle("Consolidador de Arquivos Super Poderoso")
        self.setGeometry(100, 100, 1000, 750) 

        self.current_files_paths = {}
        self.output_file_path = "" 
        self.consolidation_thread = None
        self.sheet_selections = {} # <--- NOVO: Dicionário para seleções de abas
        self.last_used_input_folder = self._load_last_input_folder() # <--- CARREGAR AO INICIAR

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- 1. Seção de Seleção de Pasta ---
        folder_selection_layout = QHBoxLayout()
        self.folder_path_label = QLabel("Pasta do Projeto:")
        self.folder_path_line_edit = QLineEdit()
        self.folder_path_line_edit.setReadOnly(True)
        # Se uma pasta foi carregada, exibi-la (opcional, ou só usar no diálogo)
        if self.last_used_input_folder:
            self.folder_path_line_edit.setText(self.last_used_input_folder)
            # Poderia até chamar list_files_in_folder aqui se quisesse carregar automaticamente
            # self.list_files_in_folder(self.last_used_input_folder) 
            # Mas vamos manter o clique do usuário para carregar os arquivos.
        self.select_folder_button = QPushButton("Selecionar Pasta...")
        self.select_folder_button.clicked.connect(self.open_folder_dialog)

        # BOTÃO para mapear cabeçalhos
        self.map_headers_button = QPushButton("Analisar/Mapear Cabeçalhos")
        self.map_headers_button.clicked.connect(self.open_header_mapping_dialog)
        self.map_headers_button.setEnabled(False) # Habilitar após selecionar pasta 
        
        folder_selection_layout.addWidget(self.folder_path_label)
        folder_selection_layout.addWidget(self.folder_path_line_edit)
        folder_selection_layout.addWidget(self.select_folder_button)
        folder_selection_layout.addWidget(self.map_headers_button)
        main_layout.addLayout(folder_selection_layout)

        # --- 2. Seção Intermediária (Arquivos/Abas e Console) ---
        middle_section_layout = QHBoxLayout()
        left_panel_layout = QVBoxLayout()
        
        self.files_label = QLabel("Arquivos Encontrados (.xlsx, .csv, .xls):")
        self.files_list_widget = QListWidget()
        self.files_list_widget.currentItemChanged.connect(self.on_file_selected_for_preview)
        
        self.sheets_label = QLabel("Abas do Arquivo Excel Selecionado (marque para incluir):")
        self.sheets_list_widget = QListWidget()
        self.sheets_list_widget.setEnabled(False) 
        # Conectar o sinal itemChanged para quando o estado de um checkbox de aba mudar
        self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)
        self.sheets_list_widget.currentItemChanged.connect(self.on_sheet_list_item_selected_for_preview)

        left_panel_layout.addWidget(self.files_label)
        left_panel_layout.addWidget(self.files_list_widget)
        left_panel_layout.addWidget(self.sheets_label)
        left_panel_layout.addWidget(self.sheets_list_widget)
        
        middle_section_layout.addLayout(left_panel_layout)

        # 2.2 Painel Direito (Abas para Console e Pré-visualização)
        right_tab_widget = QTabWidget() # <--- NOVO QTabWidget

        # Aba do Console de Log
        log_console_widget = QWidget()
        log_layout = QVBoxLayout(log_console_widget)
        self.log_label = QLabel("Console de Log:") # Pode ser removido se o título da aba for suficiente
        self.log_console_text_edit = QTextEdit()
        self.log_console_text_edit.setReadOnly(True)
        
        log_layout.addWidget(self.log_console_text_edit) # Adiciona diretamente, sem label se preferir
        right_tab_widget.addTab(log_console_widget, "Console de Log")

        # Aba de Pré-visualização de Dados
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        self.preview_table_view = QTableView()
        self.preview_table_model = PolarsTableModel() # Instancia nosso modelo customizado
        self.preview_table_view.setModel(self.preview_table_model)
        # self.preview_table_view.setEditTriggers(QTableView.NoEditTriggers) # Desabilitar edição
        self.preview_table_view.setAlternatingRowColors(True)
        preview_layout.addWidget(self.preview_table_view)
        right_tab_widget.addTab(preview_widget, "Pré-visualização de Dados")
        
        middle_section_layout.addWidget(right_tab_widget) # Adiciona o QTabWidget ao layout
        
        middle_section_layout.setStretchFactor(left_panel_layout, 1) 
        middle_section_layout.setStretchFactor(right_tab_widget, 3) # Dar mais espaço para o painel direito
        main_layout.addLayout(middle_section_layout)

        # --- 3. Seção de Configuração de Saída ---
        output_config_layout = QHBoxLayout()
        self.output_name_label = QLabel("Nome do Arquivo de Saída:")
        self.output_name_line_edit = QLineEdit("consolidado") # Nome padrão
        
        self.output_format_label = QLabel("Formato:")
        self.output_format_combo_box = QComboBox()
        self.output_format_combo_box.addItems(["XLSX", "CSV"])
        self.output_format_combo_box.currentTextChanged.connect(self.update_output_filename_extension)
        
        self.save_as_button = QPushButton("Salvar Como...")
        self.save_as_button.clicked.connect(self.open_save_file_dialog)

        output_config_layout.addWidget(self.output_name_label)
        output_config_layout.addWidget(self.output_name_line_edit)
        output_config_layout.addWidget(self.output_format_label)
        output_config_layout.addWidget(self.output_format_combo_box)
        output_config_layout.addWidget(self.save_as_button)
        main_layout.addLayout(output_config_layout)

        # --- 4. Seção de Ação e Progresso ---
        action_progress_layout = QVBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)

        self.consolidate_button = QPushButton("Iniciar Consolidação")
        self.consolidate_button.clicked.connect(self.start_consolidation) 

        self.cancel_button = QPushButton("Cancelar")
        self.cancel_button.clicked.connect(self.cancel_consolidation)
        self.cancel_button.setVisible(False) 

        buttons_layout = QHBoxLayout() 
        buttons_layout.addWidget(self.consolidate_button)
        buttons_layout.addWidget(self.cancel_button)
        
        action_progress_layout.addWidget(self.progress_bar)
        action_progress_layout.addLayout(buttons_layout) 
        main_layout.addLayout(action_progress_layout)

        main_layout.setSpacing(15)
        self.show()
        self.log_message("Aplicação iniciada. Selecione uma pasta para começar.", LogLevel.INFO)
        self.update_output_filename_extension(self.output_format_combo_box.currentText())

    def _get_config_path(self):
        """Retorna o caminho para o arquivo de configuração da aplicação."""
        # Salvar na pasta do usuário (mais robusto) ou na pasta da aplicação
        # Usar AppData ou .config no Linux/macOS é o ideal, mas para simplicidade:
        try:
            # Tenta obter o diretório do script
            base_path = os.path.dirname(os.path.abspath(sys.argv[0]))
        except:
            # Fallback para o diretório de trabalho atual se sys.argv[0] não for confiável (ex: PyInstaller one-file)
            base_path = os.getcwd() 
        return os.path.join(base_path, CONFIG_FILE_NAME)

    def _load_last_input_folder(self):
        """Carrega o último caminho da pasta de entrada do arquivo de configuração."""
        config_path = self._get_config_path()
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    return config_data.get("last_input_folder")
        except Exception as e:
            # Não precisa ser um erro crítico, apenas logar um aviso
            print(f"Aviso: Não foi possível carregar a configuração: {e}") # Usar print para log antes do logger da GUI estar pronto
        return None

    def _save_last_input_folder(self, folder_path):
        """Salva o caminho da pasta de entrada no arquivo de configuração."""
        config_path = self._get_config_path()
        config_data = {"last_input_folder": folder_path}
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=4)
            # Não precisa logar na GUI cada vez que salva, a menos que queira
            # self.log_message("Diretório padrão salvo.", LogLevel.INFO) 
        except Exception as e:
            self.log_message(f"Erro ao salvar configuração de diretório: {e}", LogLevel.ERROR)


    def open_folder_dialog(self):
        # Usar o último diretório salvo ou o diretório do QLineEdit ou o home do usuário
        start_dir = self.last_used_input_folder or self.folder_path_line_edit.text() or os.path.expanduser("~")
        
        folder_path = QFileDialog.getExistingDirectory(self, "Selecionar Pasta do Projeto", start_dir)
        
        if folder_path:
            self.folder_path_line_edit.setText(folder_path)
            self.log_message(f"Pasta selecionada: {folder_path}", LogLevel.INFO)
            self.list_files_in_folder(folder_path)
            self.last_used_input_folder = folder_path # Atualizar para o próximo uso
            self._save_last_input_folder(folder_path) # <--- SALVAR AO SELECIONAR
        else:
            self.log_message("Seleção de pasta cancelada.", LogLevel.INFO)
    
    # --- Métodos para Mapeamento de Cabeçalhos ---
    def open_header_mapping_dialog(self):
        """Coleta cabeçalhos e abre o diálogo de mapeamento."""
        if not self.current_files_paths:
            self.log_message("Selecione uma pasta e arquivos primeiro.", LogLevel.WARNING)
            return

        self.log_message("Analisando cabeçalhos dos arquivos selecionados...", LogLevel.INFO)
        unique_headers = set()
        
        # Usar get_files_and_sheets_to_process para saber quais arquivos/abas considerar
        files_and_sheets_config = self.get_files_and_sheets_to_process()
        if not files_and_sheets_config:
            self.log_message("Nenhum arquivo/aba configurado para processamento. Não é possível mapear cabeçalhos.", LogLevel.WARNING)
            return

        # Pequena thread para não travar a GUI ao ler cabeçalhos
        # (Pode ser excessivo para apenas cabeçalhos, mas bom se houver muitos arquivos)
        # Por simplicidade no MVP do mapeamento, vamos fazer síncrono por enquanto.
        # Se ficar lento, podemos mover para uma thread.
        
        processed_for_headers = 0
        for file_path, selected_sheets in files_and_sheets_config:
            file_name = os.path.basename(file_path)
            try:
                if file_path.lower().endswith(".csv"):
                    reader = pl.read_csv_batched(file_path, batch_size=5, infer_schema_length=0) 
                    batches = reader.next_batches(1) 
                    
                    if batches and len(batches) > 0:
                            first_batch = batches[0]
                            if first_batch is not None and first_batch.width > 0:
                                unique_headers.update(first_batch.columns)
                    # else: Tratar CSV vazio ou só com cabeçalho (opcional)

                elif file_path.lower().endswith((".xlsx", ".xls")) and selected_sheets:
                    for sheet_name in selected_sheets:
                        # Ler apenas a primeira linha da aba
                        # df_header = pl.read_excel(file_path, sheet_name=sheet_name, n_rows=1, infer_schema_length=0) # n_rows não é para read_excel
                        # Ler e pegar head(0).columns ou head(1).columns
                        df_sample = pl.read_excel(file_path, sheet_name=sheet_name).head(0) # Pega só cabeçalhos
                        if df_sample.width > 0:
                             unique_headers.update(df_sample.columns)
                processed_for_headers +=1
                # TODO: Adicionar feedback de progresso se for demorado
            except Exception as e:
                self.log_message(f"Erro ao ler cabeçalhos de '{file_name}' (Aba: {sheet_name if selected_sheets else 'N/A'}): {e}", LogLevel.ERROR)

        if not unique_headers:
            self.log_message("Nenhum cabeçalho encontrado ou erro ao ler todos os cabeçalhos.", LogLevel.WARNING)
            return

        self.log_message(f"Cabeçalhos únicos encontrados: {len(unique_headers)}", LogLevel.SUCCESS)
        
        # Passar o self.header_mapping existente para o diálogo
        dialog = HeaderMappingDialog(unique_headers, self, self.header_mapping)
        if dialog.exec() == QDialog.Accepted: # .exec_() em PyQt5
            self.header_mapping = dialog.get_mapping()
            self.log_message("Mapeamento de cabeçalhos atualizado.", LogLevel.SUCCESS)
            # Logar o mapeamento para depuração (opcional)
            # for original, map_info in self.header_mapping.items():
            #     self.log_message(f"  '{original}' -> '{map_info['final_name']}' (Incluir: {map_info['include']})", LogLevel.INFO)
        else:
            self.log_message("Mapeamento de cabeçalhos cancelado.", LogLevel.INFO)
    
    # --- Novos Métodos para Pré-visualização ---
    def on_file_selected_for_preview(self, current_file_item, previous_file_item):
        """Chamado quando um ARQUIVO é selecionado na lista.
           Também chama o antigo on_file_selected para carregar as abas e suas seleções.
        """
        self.on_file_selected(current_file_item, previous_file_item) # Chama a lógica existente de abas
        
        # Limpar pré-visualização se nenhum item ou arquivo não Excel/CSV
        if not current_file_item:
            self.preview_table_model.clear_data()
            return

        file_name = current_file_item.text()
        file_path = self.current_files_paths.get(file_name)

        if not file_path:
            self.preview_table_model.clear_data()
            return

        if file_path.lower().endswith(".csv"):
            self.update_preview(file_path)
        elif file_path.lower().endswith((".xlsx", ".xls")):
            # Para Excel, a pré-visualização será acionada por on_sheet_list_item_selected_for_preview
            # ou se houver apenas uma aba, podemos visualizá-la diretamente.
            # Por agora, vamos esperar a seleção de aba. Ou, se a lista de abas tiver itens,
            # pegar o primeiro item da lista de abas (se houver) e tentar visualizá-lo.
            if self.sheets_list_widget.count() > 0:
                # Tenta selecionar a primeira aba da lista para acionar a pré-visualização dela
                first_sheet_item = self.sheets_list_widget.item(0)
                if first_sheet_item: # Garante que o item existe
                    self.sheets_list_widget.setCurrentItem(first_sheet_item) 
                    # A chamada acima deve acionar on_sheet_list_item_selected_for_preview
            else: # Arquivo Excel sem abas visíveis (ou erro ao listar)
                self.preview_table_model.clear_data()
        else:
            self.preview_table_model.clear_data()


    def on_sheet_list_item_selected_for_preview(self, current_sheet_item, previous_sheet_item):
        """Chamado quando uma ABA é selecionada na lista de abas."""
        if not current_sheet_item:
            self.preview_table_model.clear_data()
            return

        current_file_item = self.files_list_widget.currentItem()
        if not current_file_item:
            self.preview_table_model.clear_data()
            return

        file_name = current_file_item.text()
        file_path = self.current_files_paths.get(file_name)
        sheet_name = current_sheet_item.text()

        if file_path and sheet_name:
            self.update_preview(file_path, sheet_name)
        else:
            self.preview_table_model.clear_data()


    def update_preview(self, file_path, sheet_name=None, n_rows_to_preview=50):
        self.log_message(f"Tentando gerar pré-visualização para: {os.path.basename(file_path)}" + (f" - Aba: {sheet_name}" if sheet_name else ""), LogLevel.INFO)
        try:
            df_preview_full = None 
            df_preview_sliced = None 

            if file_path.lower().endswith(".csv"):
                df_preview_sliced = pl.read_csv(file_path, n_rows=n_rows_to_preview, try_parse_dates=True) # try_parse_dates é válido aqui
            elif file_path.lower().endswith((".xlsx", ".xls")) and sheet_name:
                # Simplificando a leitura do Excel, a inferência de datas padrão do Polars é geralmente boa.
                df_preview_full = pl.read_excel(file_path, sheet_name=sheet_name) 
                if df_preview_full is not None:
                    df_preview_sliced = df_preview_full.head(n_rows_to_preview)
            
            if df_preview_sliced is not None and df_preview_sliced.height > 0:
                self.preview_table_model.load_data(df_preview_sliced)
                self.log_message(f"Pré-visualização gerada com {df_preview_sliced.height} linhas.", LogLevel.SUCCESS)
            elif df_preview_sliced is not None and df_preview_sliced.height == 0 :
                 self.log_message(f"O arquivo/aba {os.path.basename(file_path)} está vazio ou as primeiras {n_rows_to_preview} linhas estão vazias.", LogLevel.WARNING)
                 self.preview_table_model.clear_data()
            else:
                self.preview_table_model.clear_data()

        except Exception as e:
            self.log_message(f"Erro ao gerar pré-visualização para {os.path.basename(file_path)}: {e}", LogLevel.ERROR)
            self.preview_table_model.clear_data()

    def on_file_selected(self, current_file_item, previous_file_item):
        """Chamado quando um arquivo é selecionado na lista de arquivos."""
        self.sheets_list_widget.clear() # Limpa visualmente
        self.sheets_list_widget.setEnabled(False)

        if not current_file_item:
            return

        selected_file_name = current_file_item.text()
        file_path = self.current_files_paths.get(selected_file_name)

        if not file_path:
            self.log_message(f"Caminho não encontrado para o arquivo: {selected_file_name}", LogLevel.ERROR)
            return

        # Desconectar temporariamente o sinal para evitar chamadas recursivas ou indesejadas
        # ao popular programaticamente a lista de abas.
        try:
            self.sheets_list_widget.itemChanged.disconnect(self.on_sheet_selection_changed)
        except RuntimeError: # Se não estava conectado ainda
            pass 

        sheet_names_from_file = []
        try:
            if file_path.lower().endswith(".xlsx"):
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names_from_file = workbook.sheetnames
            elif file_path.lower().endswith(".xls"):
                workbook = xlrd.open_workbook(file_path, on_demand=True)
                sheet_names_from_file = workbook.sheet_names()
            else: # CSV ou outro tipo, sem abas
                self.log_message(f"Arquivo selecionado: {selected_file_name} (Não é Excel).", LogLevel.INFO)
                # Reconectar o sinal
                self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)
                return

            if sheet_names_from_file:
                self.sheets_list_widget.setEnabled(True)
                
                # Verificar se já temos seleções salvas para este arquivo
                # Se não, criamos uma entrada com todas as abas marcadas por padrão
                if file_path not in self.sheet_selections:
                    self.sheet_selections[file_path] = {name: True for name in sheet_names_from_file}
                
                # Obter as seleções atuais para este arquivo (pode ter sido modificado)
                current_sheet_states = self.sheet_selections[file_path]

                for sheet_name in sheet_names_from_file:
                    item = QListWidgetItem(sheet_name)
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                    # Define o estado do checkbox com base no que está salvo em self.sheet_selections
                    # Se a aba não existir mais no arquivo, mas estiver em sheet_selections (raro), ignora.
                    # Se a aba for nova no arquivo, o default é True (do passo anterior).
                    is_checked = current_sheet_states.get(sheet_name, True) # Default para True se for nova
                    item.setCheckState(Qt.Checked if is_checked else Qt.Unchecked)
                    self.sheets_list_widget.addItem(item)
                
                log_suffix = " (carregadas do cache)" if file_path in self.sheet_selections and any(not s for s in current_sheet_states.values()) else " (todas marcadas por padrão)"
                self.log_message(f"Abas encontradas para '{selected_file_name}': {', '.join(sheet_names_from_file)}.{log_suffix}", LogLevel.INFO)

            else:
                self.log_message(f"Nenhuma aba encontrada no arquivo Excel: {selected_file_name}", LogLevel.WARNING)
                # Se não há abas, limpar qualquer entrada antiga em sheet_selections
                if file_path in self.sheet_selections:
                    del self.sheet_selections[file_path]
        
        except Exception as e:
            self.log_message(f"Erro ao ler abas do arquivo {selected_file_name}: {e}", LogLevel.ERROR)
            if file_path in self.sheet_selections: # Limpar se deu erro
                 del self.sheet_selections[file_path]
        finally:
            # Reconectar o sinal itemChanged após a população
            self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)


    def on_sheet_selection_changed(self, item_changed: QListWidgetItem):
        """Chamado quando o estado de um checkbox de aba muda."""
        current_file_item = self.files_list_widget.currentItem()
        if not current_file_item:
            return # Nenhum arquivo Excel selecionado na lista principal

        selected_file_name = current_file_item.text()
        file_path = self.current_files_paths.get(selected_file_name)

        if not file_path or not file_path.lower().endswith((".xlsx", ".xls")):
            return # Não é um arquivo Excel válido ou caminho não encontrado

        sheet_name = item_changed.text()
        is_checked = item_changed.checkState() == Qt.Checked

        # Garantir que a entrada para o arquivo exista em self.sheet_selections
        if file_path not in self.sheet_selections:
            # Isso não deveria acontecer se on_file_selected populou corretamente,
            # mas é uma salvaguarda.
            self.sheet_selections[file_path] = {} 
        
        self.sheet_selections[file_path][sheet_name] = is_checked
        # self.log_message(f"Seleção da aba '{sheet_name}' para '{selected_file_name}' atualizada para: {'Marcada' if is_checked else 'Desmarcada'}", LogLevel.INFO)

    def get_files_and_sheets_to_process(self):
        files_to_process_list = []
        if not self.current_files_paths:
             self.log_message("Nenhuma pasta selecionada ou nenhum arquivo encontrado.", LogLevel.WARNING)
             return None

        for file_name, file_path in self.current_files_paths.items():
            selected_sheets_for_file = []
            is_excel = file_path.lower().endswith((".xlsx", ".xls"))

            if is_excel:
                # Usar as seleções armazenadas em self.sheet_selections
                if file_path in self.sheet_selections:
                    for sheet_name, is_selected in self.sheet_selections[file_path].items():
                        if is_selected:
                            selected_sheets_for_file.append(sheet_name)
                    
                    if not selected_sheets_for_file:
                        self.log_message(f"Nenhuma aba marcada para o arquivo Excel '{file_name}' nas configurações. Será pulado.", LogLevel.WARNING)
                        # Não adiciona à lista de processamento se explicitamente nenhuma aba foi marcada
                        continue # Pula para o próximo arquivo
                else:
                    # Fallback: Se por algum motivo não há entrada (ex: arquivo adicionado e nunca selecionado),
                    # processar todas as abas como padrão.
                    self.log_message(f"Nenhuma seleção de abas encontrada para '{file_name}'. Processando todas as abas por padrão.", LogLevel.INFO)
                    try:
                        if file_path.lower().endswith(".xlsx"):
                            workbook = openpyxl.load_workbook(file_path, read_only=True)
                            selected_sheets_for_file = workbook.sheetnames
                        elif file_path.lower().endswith(".xls"):
                            workbook = xlrd.open_workbook(file_path, on_demand=True)
                            selected_sheets_for_file = workbook.sheet_names()
                    except Exception as e:
                        self.log_message(f"Erro ao tentar obter todas as abas de '{file_name}' para fallback: {e}. Pulando arquivo.", LogLevel.ERROR)
                        continue # Pula este arquivo
                
                if not selected_sheets_for_file: # Double check, caso o fallback não encontre abas
                    self.log_message(f"Arquivo Excel '{file_name}' não tem abas (ou erro ao listá-las no fallback). Será pulado.", LogLevel.WARNING)
                    continue

                files_to_process_list.append((file_path, selected_sheets_for_file))

            elif file_path.lower().endswith(".csv"):
                files_to_process_list.append((file_path, None)) # None indica CSV
            else:
                self.log_message(f"Arquivo '{file_name}' não é Excel nem CSV. Pulando.", LogLevel.WARNING)
        
        if not files_to_process_list:
            self.log_message("Nenhum arquivo encontrado ou selecionado para processamento após filtrar.", LogLevel.WARNING)
            return None
            
        return files_to_process_list


    def log_message(self, message, level=LogLevel.INFO):
        self.log_console_text_edit.append(f"{level.value} {message}")

    def list_files_in_folder(self, folder_path):
        self.files_list_widget.clear()
        self.sheet_selections.clear() 
        self.sheets_list_widget.clear()
        self.sheets_list_widget.setEnabled(False)
        self.current_files_paths.clear()
        self.preview_table_model.clear_data() 
        
        # Desabilitar o botão no início da função ou quando não há arquivos
        self.map_headers_button.setEnabled(False) 
        # Limpar mapeamento de cabeçalhos se a pasta mudar
        self.header_mapping.clear()


        supported_extensions = ("*.xlsx", "*.csv", "*.xls")
        found_files_paths = []
        try:
            for ext in supported_extensions:
                pattern = os.path.join(folder_path, ext)
                found_files_paths.extend(glob.glob(pattern))
            
            if found_files_paths:
                for full_path in found_files_paths:
                    file_name = os.path.basename(full_path)
                    self.files_list_widget.addItem(file_name)
                    self.current_files_paths[file_name] = full_path
                self.log_message(f"Encontrados {len(found_files_paths)} arquivos na pasta.", LogLevel.SUCCESS)
                self.map_headers_button.setEnabled(True) # HABILITAR AQUI se arquivos forem encontrados
            else:
                self.log_message("Nenhum arquivo suportado (.xlsx, .csv, .xls) encontrado na pasta.", LogLevel.WARNING)
                # self.map_headers_button.setEnabled(False) # Já está desabilitado pelo início da função

        except Exception as e:
            self.log_message(f"Erro ao listar arquivos: {e}", LogLevel.ERROR)
            self.current_files_paths.clear()
            # self.map_headers_button.setEnabled(False) # Já está desabilitado
    
    def update_output_filename_extension(self, selected_format):
        current_name = self.output_name_line_edit.text()
        name_part, _ = os.path.splitext(current_name)
        new_extension = "." + selected_format.lower()
        self.output_name_line_edit.setText(name_part + new_extension)

    def open_save_file_dialog(self):
        current_folder = self.folder_path_line_edit.text() or os.path.expanduser("~")
        suggested_filename = self.output_name_line_edit.text()
        initial_path = os.path.join(current_folder, suggested_filename)
        selected_format = self.output_format_combo_box.currentText()
        if selected_format == "XLSX":
            filter_str = "Arquivos Excel (*.xlsx)"
        elif selected_format == "CSV":
            filter_str = "Arquivos CSV (*.csv)"
        else:
            filter_str = "Todos os Arquivos (*)"
        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar Arquivo Consolidado Como...", initial_path, filter_str)
        if file_path:
            self.output_file_path = file_path
            base_name = os.path.basename(file_path)
            self.output_name_line_edit.setText(base_name)
            name_part, ext_part = os.path.splitext(base_name)
            if ext_part.lower() == ".xlsx" and self.output_format_combo_box.currentText() != "XLSX":
                self.output_format_combo_box.setCurrentText("XLSX")
            elif ext_part.lower() == ".csv" and self.output_format_combo_box.currentText() != "CSV":
                self.output_format_combo_box.setCurrentText("CSV")
            self.log_message(f"Arquivo de saída definido como: {file_path}", LogLevel.SUCCESS)
        else:
            self.log_message("Seleção de local para salvar cancelada.", LogLevel.INFO)
    
    def start_consolidation(self):
        if not self.folder_path_line_edit.text():
            self.log_message("Por favor, selecione uma pasta de projeto primeiro.", LogLevel.WARNING)
            return

        if not self.output_file_path:
            self.log_message("Por favor, defina o arquivo de saída usando 'Salvar Como...'.", LogLevel.WARNING)
            return

        files_to_process = self.get_files_and_sheets_to_process()
        if not files_to_process: # Checa se é None ou lista vazia
            self.log_message("Nenhum arquivo ou aba válida selecionada para consolidação.", LogLevel.WARNING)
            return

        self.set_ui_for_processing(True)
        output_format = self.output_format_combo_box.currentText()
        
        self.consolidation_thread = ConsolidationWorker(files_to_process, self.output_file_path, output_format, self.header_mapping)
        self.consolidation_thread.log_message.connect(self.log_message) 
        self.consolidation_thread.progress_updated.connect(self.update_progress_bar)
        self.consolidation_thread.finished.connect(self.on_consolidation_finished)
        
        self.consolidation_thread.start()

    def update_progress_bar(self, value):
        self.progress_bar.setValue(value)

    def on_consolidation_finished(self, success, message):
        if success:
            self.log_message(f"Resultado: {message}", LogLevel.SUCCESS)
        else:
            self.log_message(f"Resultado: {message}", LogLevel.ERROR)
        
        self.set_ui_for_processing(False)
        self.consolidation_thread = None 

    def set_ui_for_processing(self, processing):
        not_proc = not processing
        self.select_folder_button.setEnabled(not_proc)
        self.files_list_widget.setEnabled(not_proc)
        
        # Habilitar o botão de mapear apenas se não estiver processando E houver arquivos listados
        self.map_headers_button.setEnabled(not_proc and self.files_list_widget.count() > 0) 

        has_sel_excel_sheets = False
        current_file_item = self.files_list_widget.currentItem()
        if not_proc and current_file_item: # Só verifica se não estiver processando e houver item
            fn = current_file_item.text()
            fp = self.current_files_paths.get(fn)
            if fp and (fp.lower().endswith((".xlsx",".xls"))) and self.sheets_list_widget.count() > 0:
                has_sel_excel_sheets = True
        self.sheets_list_widget.setEnabled(not_proc and has_sel_excel_sheets)
        
        self.output_name_line_edit.setEnabled(not_proc)
        self.output_format_combo_box.setEnabled(not_proc)
        self.save_as_button.setEnabled(not_proc)
        self.consolidate_button.setVisible(not_proc) 
        self.cancel_button.setVisible(processing)       
        if not_proc:
            self.progress_bar.setValue(0)

    def cancel_consolidation(self):
        if self.consolidation_thread and self.consolidation_thread.isRunning():
            self.log_message("Solicitando cancelamento da consolidação...", LogLevel.INFO)
            self.consolidation_thread.stop() 
        else:
            self.log_message("Nenhuma consolidação em andamento para cancelar.", LogLevel.INFO)
            self.set_ui_for_processing(False) # Garante que a UI volte ao normal se o botão for clicado por engano

    def open_folder_dialog(self): # Adicionado para garantir que está presente
        start_dir = self.folder_path_line_edit.text() or os.path.expanduser("~")
        folder_path = QFileDialog.getExistingDirectory(self, "Selecionar Pasta do Projeto", start_dir)
        if folder_path:
            self.folder_path_line_edit.setText(folder_path)
            self.log_message(f"Pasta selecionada: {folder_path}", LogLevel.INFO)
            self.list_files_in_folder(folder_path)
        else:
            self.log_message("Seleção de pasta cancelada.", LogLevel.INFO)

    def closeEvent(self, event):
        if self.consolidation_thread and self.consolidation_thread.isRunning():
            self.log_message("Fechando aplicação, parando consolidação em andamento...", LogLevel.INFO)
            self.consolidation_thread.stop()
            self.consolidation_thread.wait() 
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec())